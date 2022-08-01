programName = 'create_opcua_client_2'

# Create Test Client Machine 2:
from asyncio.windows_events import NULL
from distutils.command.build import build
from platform import node
from re import S
from openpyxl import workbook
from openpyxl.workbook import Workbook
from openpyxl.reader.excel import load_workbook
import openpyxl
from opcua import Client, ua
import time
from opcua.ua.uaprotocol_auto import VariableAttributes
from opcua.ua.uaerrors import UaError, BadTimeout, BadNoSubscription, BadSessionClosed
from opcua.common.connection import SecureConnection
node_list_1 = []
node_list_2 = []
node_list_3 = []
node_list = []
i = 0
#-------------------------------------------------------------
#CONNECTING TO AN OPC-UA SERVER
class Connection:
    def __init__(self,url):
        self.url = url
        self.client = Client(self.url)
        self.root = self.client.get_root_node()
    def set_usr_pwd(self,user, password):
        self.client.set_user(user)
        self.client.set_password(password)
        self.client.connect()
        print("Client Connected to server OK")
    def disconnect(self):
        self.client.disconnect()
def build_opc_tree_1(node,client):
        nodeClass = node.get_node_class()
        children  = node.get_children()
        for child in children:
            child_ref = client.get_node(child)
            childClass = child_ref.get_node_class()
            node_list.append([node, nodeClass, child_ref, childClass])
            if childClass == ua.NodeClass.Object:
                build_opc_tree_1(child_ref,client)
        return node_list

class write: 
    def __init__(self,filepath):
        self.indexes = []
        self.filepath = filepath
    def writing_excel(self,client,node_list):
        self.wb = Workbook()
        #filepath = "C:/Users/jimmy.carradore/Documents/GitHub/OPC-UA-project/sample.xlsx"
        self.wb = load_workbook(self.filepath)
        self.sheet = self.wb.active
        i = 1
        for j in range(0 ,len(node_list)-1):
            if '.' in str(client.get_node(node_list[j][2])):
                list = str(client.get_node(node_list[j][2])).split(".")
                #debug
                if 'AW24-T4' in list[0]:
                    if ('ALARM' in list[1]) or  ('OPERATOR'in list[1]):
                        #print(list[1])
                        continue    
                    else:
                        self.indexes.append(j)
                    self.sheet.cell(row = 1 , column = i).value = str(list[1])
                    i = i + 1
                    self.wb.save(self.filepath)
                    self.wb.close()
    def write_values(self,client,filepath,node_list,how_many):
        for i in range(2,how_many):
            counter = 1
            for j in self.indexes:
                node_ex = client.get_node(node_list[j][2])
                sensor_value = node_ex.get_value()
                self.sheet.cell(row = i , column = counter).value = sensor_value
                counter = counter + 1
        self.wb.save(filepath)
        self.wb.close()

class Sheet_ex:
    def __init__(self,filepath,rows):
        self.filepath = filepath
        self.rows = rows
    def load_sheet(self):
        self.wb = Workbook()
        #filepath = "C:/Users/jimmy.carradore/Documents/GitHub/OPC-UA-project/sample.xlsx"
        self.wb = load_workbook(self.filepath)
        self.sheet = self.wb.active
        self.counter = 1
    def clean(self): 
        self.load_sheet()
        for i in range(1 , 1000):
            for j in range (1, self.rows + 1): 
                if self.sheet.cell(row = j , column = i).value == '':
                    return
                else: self.sheet.cell(row = j , column = i).value = '' 
    def save(self):
        self.wb.save(self.filepath) 
        self.wb.close()
#-----------------------------------------------------------------------------
#HANDLE PASSWORD AND USER
#--------------------------------------------------------------------------
#try:
sheet_1 = Sheet_ex("C:/Users/jimmy.carradore/Documents/GitHub/OPC-UA-project/sample.xlsx",5)
sheet_1.load_sheet()
sheet_1.clean()
sheet_1.save()
connection1 = Connection("opc.tcp://192.168.0.53:59611")
connection1.set_usr_pwd("admin", "admin")
node_list_1 = build_opc_tree_1(connection1.root,connection1.client)
prova = write("C:/Users/jimmy.carradore/Documents/GitHub/OPC-UA-project/sample.xlsx")
prova.writing_excel(connection1.client,node_list_1)
sheet_1.load_sheet()
prova.write_values(connection1.client,"C:/Users/jimmy.carradore/Documents/GitHub/OPC-UA-project/sample.xlsx",node_list_1,6)
#sheet_1.save()
connection1.disconnect()
quit()

#--------------------------------------------------------------------------


    # while True:
    #     counter= 1
    #     s = 1
        # print('#------------------------------------#')
        # print('#          starting while cycle      #')
        # print('#------------------------------------#')


        

# except ua.utils.SocketClosedException as ex:
#     print('Exception ua.utils.SocketClosedExceptionoccured while trying to open client session')
#     print('Exception occured while trying to open client session')

# except TimeoutError as ex:
#     print('Exception timeout error')
#     print('Exception is:', ex)

# except (UaError,  BadTimeout, BadNoSubscription, BadSessionClosed) as ex:
#     print('Protocol error')
#     print('Exception is:', ex)

# except Exception as ex:
#     print('Generic Exception occured')
#     print('Exception is: ', ex)
# #-----------------------------------------------------------------------------------------------------------------------------------#
