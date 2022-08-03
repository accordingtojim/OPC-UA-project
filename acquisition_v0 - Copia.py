from asyncio.windows_events import NULL
from distutils.command.build import build
from pickle import FALSE, TRUE
from platform import node
from re import S
from openpyxl import workbook
from openpyxl.workbook import Workbook
from openpyxl.reader.excel import load_workbook
import openpyxl
from opcua import Client, ua
import time
from opcua.ua.uaprotocol_auto import VariableAttributes
from opcua.ua.uaerrors import UaError, BadTimeout, BadNoSubscription, BadSessionClosed
from opcua.common.connection import SecureConnection

#giorgio = ''.join(str(i) for i in array)
filepath = "C:/Users/jimmy.carradore/Documents/GitHub/OPC-UA-project/Cartel1.xlsx"
wb = Workbook()
wb = load_workbook(filepath)
sheet = wb.active
#sheet.cell(row = 1 , column = i).value = str(list[1])
#i = i + 1
counter = 1

flag = FALSE
matrix = []
for j in range (1,550):
    flag = FALSE
    if j < 10:
        lista = ["Y",0,0,0]
        lista.append(j)
        lista.append(".")    
        for l in range (0,8):
            if flag == FALSE:
                lista.append(l)
                flag = TRUE
                lista1 = ''.join(str(s) for s in lista)
                sheet.cell(row = counter , column = 6).value = lista1
                counter = counter + 1
            else:
                lista[6] = l
                lista1 = ''.join(str(s) for s in lista)
                sheet.cell(row = counter , column = 6).value = lista1
                counter = counter + 1
    if ((j > 9) & (j < 100)):
        lista = ["Y",0,0]
        lista.append(j)
        lista.append(".")
        for l in range (0,8):
            if flag == FALSE:
                lista.append(l)
                flag = TRUE
                lista1 = ''.join(str(s) for s in lista)
                sheet.cell(row = counter , column = 6).value = lista1
                counter = counter + 1
            else:
                lista[5] = l
                lista1 = ''.join(str(s) for s in lista)
                sheet.cell(row = counter , column = 6).value = lista1
                counter = counter + 1
    if ((j > 99) & (j < 550)):
        lista = ["Y",0]
        lista.append(j)
        lista.append(".")
        for l in range (0,8):
            if flag == FALSE:
                lista.append(l)
                flag = TRUE
                lista1 = ''.join(str(s) for s in lista)
                sheet.cell(row = counter , column = 6).value = lista1
                counter = counter + 1
            else:
                lista[4] = l
                lista1 = ''.join(str(s) for s in lista)
                sheet.cell(row = counter , column = 6).value = lista1
                counter = counter + 1           
wb.save(filepath)
wb.close()
    
        

# for j in range(0,8):
    
#     print(j)