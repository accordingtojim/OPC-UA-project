programName = 'create_opcua_client_2'

# Create Test Client Machine 2:
from re import S
from openpyxl import workbook
from openpyxl.workbook import Workbook
from openpyxl.reader.excel import load_workbook
import openpyxl
from opcua import Client, ua
import time

from opcua.ua.uaprotocol_auto import VariableAttributes

from opcua.ua.uaerrors import UaError, BadTimeout, BadNoSubscription, BadSessionClosed
from opcua.common.connection import SecureConnection
#from my_library import *
#-----------------------------------------------------------#
# connecting to am OPC-UA server
#-----------------------------------------------------------#

url = "opc.tcp://192.168.0.53:59611"
# url = "opc.tcp://192.168.100.96:4840"
client = Client(url)

#--------------------------------------------------------------------

#EXCEL SHEET MANAGEMENT
wb = Workbook()
filepath = "C:/Users/jimmy.carradore/Desktop/test/sample.xlsx"
wb = load_workbook(filepath)
sheet = wb.active
counter = 1

#--------------------------------------------------------------------------

root = client.get_root_node()
print("Root node is: ", str(root))
def build_opc_tree(node):
        nodeClass = node.get_node_class()
       # print('\n#--------------------------------------------------------------#')
        print('\nnode', node, 'has class', nodeClass)

        children  = node.get_children()
        print('\tchildren of node', node, '=', len(children))
        print('\t', children)
        h = 0
        for child in children:
            child_ref = client.get_node(child)
            childClass = child_ref.get_node_class()
            print('\t', h, '- child', child_ref, 'has class', childClass)
            node_list.append([node, nodeClass, child_ref, childClass])
            h = h + 1
            if childClass == ua.NodeClass.Object:
                build_opc_tree(child_ref)

        return
#--------------------------------------------------------------------------
 
flag_string = False
i = 0
s = 2
try:
    client.set_user("admin")
    client.set_password("admin")
    client.connect()
    print("Client Connected to server OK")

    # client.connect_socket()
    # cfs = client.connect_and_find_servers()
    # print("Client Socket Connected to server OK")
    # print("we got:", cfs)

    # uri = "FANUCOPC.SERVER.0"
    # idx = client.get_namespace_index(uri)
    # print('idx = ', idx)

    # client.create_session()
    # print("Client Session created")

    # client.activate_session(username='admin', password='admin', certificate=None)
    # print("Client Session activated")
    
   # children = node.get_children()
    #print("Children node are: ",str(children))
    #print("Children node are: ",children[1])
    #child_ref = client.get_node(children[1])
    #print("Child ref is : ",child_ref)
    #childClass = child_ref.get_node_class()
    #print("Child Class is : ",childClass)
    #print(str(childClass==ua.NodeClass.Object))


#------------------------------------------------------------------------

    
#-----------------------------------------------------------------------


    while True:
        counter=counter+1
        # print('#------------------------------------#')
        # print('#          starting while cycle      #')
        # print('#------------------------------------#')


        node_list = []
        node_dict = {}
        iter = 0

        build_opc_tree(root)
        #print("NODE LIST IS : ", node_list)
        #print("NODE LIST 2 : ", str(node_list[2]))
        #node_ex=set()
    
        # for iter in node_list:
        #     #print('\n' , i , iter)
        #     i = i + 1
        print('##############################################################################################################################')
        print("node list :", node_list )
        if flag_string == False:
            for j in range(55 , 213):
            # if flag_string == True:
            # node_ex = client.get_node(node_list[j][2])
            # sensor_value = node_ex.get_value()
            # sheet.cell(row = (j-53) , column = s).value = sensor_value
                i = i + 1
                #print(str(node_list [100] [2]))
                list = str(client.get_node(node_list[j][2])).split(".")
                list_deb = str(list [1])
                sheet.cell(row = 1 , column = i).value = list_deb
                #print(list_deb)
        flag_string == True
        wb.save(filepath) 
#---------------------------start debug----------------------------------------#

        # node_ex = client.get_node(node_list[182][2])
        # sensor_value = node_ex.get_value()
        # sheet.cell(row = 1 , column = (182)).value = sensor_value
        # wb.save(filepath)
#---------------------------end debug------------------------------------------#
        for j in range(55 , 213):
            if j == 181 or j ==182 :
                continue
            else:
                node_ex = client.get_node(node_list[j][2])
                sensor_value = node_ex.get_value()
                sheet.cell(row = s , column = (j-54)).value = sensor_value
                wb.save(filepath)
                #print(sensor_value)
        s = s + 1

        
        #flag_string = True
           
        
        

        #print("node ex is:",str(node_ex))
        #set(node_ex)
        #print("Node value is :",node_ex.get_value())
        # getting root node of connected OPC server
       

        #objects = client.get_objects_node()
        #print("Objects node is: ", objects)


        #root_al = root.get_access_level()
        # print('\nroot access level = ', root_al)

        #ch = root.get_children()
        #print('\nroot children = ', ch)

        #node = client.get_node("ns=2;s=AW24-T4")
        #myNodeChild = node.get_children()
        #print("myNodeChild is: ", myNodeChild)
        #print("number of child nodes: ", len(myNodeChild))

        #time.sleep(1)
        

except ua.utils.SocketClosedException as ex:
    print('Exception ua.utils.SocketClosedExceptionoccured while trying to open client session')
    print('Exception occured while trying to open client session')

except TimeoutError as ex:
    print('Exception timeout error')
    print('Exception is:', ex)

except (UaError,  BadTimeout, BadNoSubscription, BadSessionClosed) as ex:
    print('Protocol error')
    print('Exception is:', ex)

except Exception as ex:
    print('Generic Exception occured')
    print('Exception is: ', ex)
#-----------------------------------------------------------------------------------------------------------------------------------#
