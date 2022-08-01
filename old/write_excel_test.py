from openpyxl import load_workbook
from openpyxl import Workbook
wb = load_workbook('test.xlsx')    #you must be in the same folder of the excel file and you must specify the file name
ws = wb.active
ws.cell(row=1,column=1).value = 'test'
#wb.save('test.xlsx')