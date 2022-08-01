programName = 'create_opcua_client_2'

# Create Test Client Machine 2:
from re import S
from openpyxl import workbook
from openpyxl.workbook import Workbook
from openpyxl.reader.excel import load_workbook
import openpyxl

from opcua import Client, ua
import time

from opcua.ua.uaprotocol_auto import VariableAttributes

from opcua.ua.uaerrors import UaError, BadTimeout, BadNoSubscription, BadSessionClosed
from opcua.common.connection import SecureConnection

#-----------------------------------------------------------#
# connecting to am OPC-UA server
#-----------------------------------------------------------#
url = "opc.tcp://192.168.0.53:59611"
# url = "opc.tcp://192.168.100.96:4840"
client = Client(url)

#--------------------------------------------------------------------

#EXCEL SHEET MANAGEMENT
wb = Workbook()
filepath = "C:/Users/jimmy.carradore/Desktop/test/sample.xlsx"
wb = load_workbook(filepath)
sheet = wb.active
counter = 2

#--------------------------------------------------------------------------

try:
    client.set_user("admin")
    client.set_password("admin")
    client.connect()
    print("Client Connected to server OK")
    

    # client.connect_socket()
    # cfs = client.connect_and_find_servers()
    # print("Client Socket Connected to server OK")
    # print("we got:", cfs)

    # uri = "FANUCOPC.SERVER.0"
    # idx = client.get_namespace_index(uri)
    # print('idx = ', idx)

    # client.create_session()
    # print("Client Session created")

    # client.activate_session(username='admin', password='admin', certificate=None)
    # print("Client Session activated")

    while True:
        #print('#------------------------------------#')
        #print('#          starting while cycle      #')
        #print('#------------------------------------#')
        # getting root node of connected OPC server
        root = client.get_root_node()
        
        objects = client.get_objects_node()     

        node_name0 = "ns=2;s=AW24-T4.CN_MACHINING_START"
        node0 = client.get_node(node_name0)
        sensor_value0 = node0.get_value()
        sheet.cell(row = counter, column = 1).value = sensor_value0

        node_name1 = "ns=2;s=AW24-T4.CN_MACHINING_STOP"
        node1 = client.get_node(node_name1)
        sensor_value1 = node1.get_value()
        sheet.cell(row = counter, column = 2).value = sensor_value1
        
        wb.save(filepath)
        counter = counter +1
        #----------------------start debug--------------------------#

        # print("Root node is: ", str(root))
        # print("Objects node is: ", objects)
        # print('\nparent node =', node)
        # print('\nnode =', node)
        # print('val =', sensor_value)

        #----------------------end debug---------------------------#

        #root_al = root.get_access_level()
        # print('\nroot access level = ', root_al)

        # ch = root.get_children()
        # print('\nroot children = ', ch)

        # node = client.get_node("ns=2;s=PMC_PULSECODE_TEMP_X1")
        # myNodeChild = node.get_children()
        # print("myNodeChild is: ", myNodeChild)
        # print("number of child nodes: ", len(myNodeChild))

        # time.sleep(1)

except ua.utils.SocketClosedException as ex:
    print('Exception ua.utils.SocketClosedExceptionoccured while trying to open client session')
    print('Exception occured while trying to open client session')

except TimeoutError as ex:
    print('Exception timeout error')
    print('Exception is:', ex)

except (UaError,  BadTimeout, BadNoSubscription, BadSessionClosed) as ex:
    print('Protocol error')
    print('Exception is:', ex)

except Exception as ex:
    print('Generic Exception occured')
    print('Exception is: ', ex)
#-----------------------------------------------------------------------------------------------------------------------------------#
