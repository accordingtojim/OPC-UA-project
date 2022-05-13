####    test to check how to modify a content of a cell
#cell_range = ws['A1':'C2']
#c = ws['G2']
#d = ws['A1']
#print(wb.sheetnames)
#if "Contact sensor" == c.value:
#    print("true")
#print(ws.cell(4,2,1).value)
#ws['A2'] = 32
#print(ws['A2'].value)
####     end test

####    test to pass a string as parameter
# def pass_string(prova):
# if prova == 'si':
#   print('true')
# pass_string('si')
####     end test
######################################################################################################################
#      THIS IS A SOFTWARE TO REPLACE CONTENTS OF CELL IN EXCEL TABLES

### must passed a row where the first element is non Null 
from openpyxl import load_workbook
from openpyxl import Workbook
wb = load_workbook('test.xlsx')    #you must be in the same folder of the excel file and you must specify the file name
ws = wb.active
def change_name(to_be_replaced, replace_with, which_row, which_column, condition):
 number_column=0
 number_row=0
 if condition == 'row':
    #print('ok')            DEBUG
    for row in ws.iter_rows(min_row=which_row, max_row=which_row, min_col=1):   #counting the columns
        #print('ok')            DEBUG
        for cell in row:
            if cell.value != None:
                number_column=number_column+1
                #print(number_column)              DEBUG
    for row in ws.iter_rows( min_row=which_row, min_col=1, max_row=which_row , max_col=number_column):      #replacing the strings
        for cell in row:
            if cell.value == to_be_replaced:          #clue part
                 cell.value =replace_with 
 elif condition == 'column':
     for col in ws.iter_cols(min_row=1,min_col=which_column, max_col=which_column):   #counting the rows
        for cell in col:
            if cell.value != None:
                number_row= number_row + 1
     for col in ws.iter_cols( min_row=1, min_col=which_column, max_col=which_column , max_row=number_row):      #replacing the strings
        for cell in col:
            if cell.value == to_be_replaced:                 #clue part
                cell.value =replace_with 
 else:
     print('error: you must write "row" or "column" as fifth parameter')
 wb.save('test.xlsx')

#change_name('temperature','cutto',5 ,0 , 'row')







      





