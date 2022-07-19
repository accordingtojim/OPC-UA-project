# Create Test Client Machine 2:
from asyncio.windows_events import NULL
from platform import node
from re import S
from openpyxl import workbook
from openpyxl.workbook import Workbook
from openpyxl.reader.excel import load_workbook
import openpyxl
from opcua import Client, ua
import time
from opcua.ua.uaprotocol_auto import VariableAttributes
from opcua.ua.uaerrors import UaError, BadTimeout, BadNoSubscription, BadSessionClosed
from opcua.common.connection import SecureConnection
node_list = []
class myclass:
    def __init__(self,test):
        self.test = test
    def func(self):
        self.test2= self.test + 2
        return self.test2
variable = myclass(0)  
print(variable.func())

class Sheet_ex:
    def __init__(self,filepath,rows):
        self.filepath = filepath
        self.rows = rows
        self.load_sheet()
        self.clean()
        self.save()
    def load_sheet(self):
        self.wb = Workbook()
        #filepath = "C:/Users/jimmy.carradore/Documents/GitHub/OPC-UA-project/sample.xlsx"
        self.wb = load_workbook(self.filepath)
        self.sheet = self.wb.active
        self.counter = 1
    def clean(self): 
        self.load_sheet()
        for i in range(1 , 1000):
            for j in range (1, self.rows + 1): 
                if self.sheet.cell(row = j , column = i).value == '':
                    return
                else: self.sheet.cell(row = j , column = i).value = ''
    def save(self): 
        self.wb.save(self.filepath) 
        self.wb.close()
#ws = wb['test1']

sheet_1 = Sheet_ex("C:/Users/jimmy.carradore/Documents/GitHub/OPC-UA-project/sample.xlsx",2)



